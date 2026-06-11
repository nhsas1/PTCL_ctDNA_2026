#!/usr/bin/env python3
# ============================================================
# PTCL ctDNA — Excel QC Master Sheet Update — Batch 3
# Author: Noor Shaban
# Date: June 2026
# Purpose: Add 13 Batch 3 samples to PTCL_ctDNA_QC_Master_v2.xlsx
# ============================================================

import openpyxl
from copy import copy

EXCEL_PATH = '/scratch/alice/n/nhsas1/ptcl_ctdna_thesis/QC_results/PTCL_ctDNA_QC_Master_v2.xlsx'
OUTPUT_PATH = '/scratch/alice/n/nhsas1/ptcl_ctdna_thesis/QC_results/PTCL_ctDNA_QC_Master_v3.xlsx'

# ============================================================
# BATCH 3 DATA — all 13 samples
# Columns match exactly:
# QC_Summary: SampleID, Filename, Batch, SeqDate, ProcDate,
#             TotalReads, MappedReads, MappingPct, DupReads,
#             DupPct, ReadLen, CoverageEst, MeanDepth, RefGenome,
#             RGTag, BQSR, DupTool, DupStatus, BAMIndex,
#             Sex, MeanBaseQ, Notes
# ============================================================

qc_summary_rows = [
    ['S22', 'LB_20241212_1_SI.bam', 'B3', 'Dec 2024', 'Dec 2024',
     24812836, 24788902, '99.90%', 0, '0%', '150bp', '~1.16x', '1.084x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.5', None],

    ['S23', 'LB_20241212_2_SI.bam', 'B3', 'Dec 2024', 'Dec 2024',
     25777613, 25757022, '99.92%', 0, '0%', '150bp', '~1.20x', '1.084x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.5', None],

    ['S24', 'LB_20241212_3_SI.bam', 'B3', 'Dec 2024', 'Dec 2024',
     23441384, 23416051, '99.89%', 0, '0%', '150bp', '~1.09x', '1.084x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.5', None],

    ['S25', 'LB_20241212_4_SI.bam', 'B3', 'Dec 2024', 'Dec 2024',
     26727813, 26700517, '99.90%', 0, '0%', '150bp', '~1.25x', '1.084x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'Female', '~39.5', None],

    ['S26', 'LB_20241212_6_SI.bam', 'B3', 'Dec 2024', 'Dec 2024',
     23185841, 23158732, '99.88%', 0, '0%', '150bp', '~1.08x', '1.084x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.5', 'chr13/14/15/22 borderline low — monitor'],

    ['S27', 'LB_241114_1_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     31417067, 31395810, '99.93%', 0, '0%', '150bp', '~1.47x', '1.390x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.1', 'chr2 and chr5 elevated — flag for QDNAseq'],

    ['S28', 'LB_241114_2_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     37772420, 37742429, '99.92%', 0, '0%', '150bp', '~1.77x', '1.770x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Female', '~39.1', 'chr2 elevated — flag for QDNAseq'],

    ['S29', 'LB_241114_3_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     33979291, 33956341, '99.93%', 0, '0%', '150bp', '~1.59x', '1.590x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Female', '~39.1', 'HIGH PRIORITY: chr6/10/12 reduced chr18 elevated'],

    ['S30', 'LB_241114_4_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     33019893, 32997791, '99.93%', 0, '0%', '150bp', '~1.54x', '1.540x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Female', '~39.1', 'HIGH PRIORITY: chr18 strongly elevated +49%'],

    ['S31', 'LB_241114_5_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     30708479, 30685391, '99.92%', 0, '0%', '150bp', '~1.43x', '1.430x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.0', 'chr2 mild elevation — flag for QDNAseq'],

    ['S32', 'LB_241114_6_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     35861737, 35829932, '99.91%', 0, '0%', '150bp', '~1.68x', '1.680x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.1', 'HIGH PRIORITY: chr17/19/20 all elevated'],

    ['S33', 'LB_241114_7_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     31603742, 31578421, '99.92%', 0, '0%', '150bp', '~1.48x', '1.480x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Male', '~39.0', 'chr19 and chr20 mild elevation — flag for QDNAseq'],

    ['S34', 'LB_241114_8_SI.bam', 'B3', 'Nov 2024', 'Nov 2024',
     33741119, 33703139, '99.89%', 0, '0%', '150bp', '~1.58x', '1.580x',
     'hg38', 'NO', 'NOT_CONFIRMED', 'Not run', 'Not run', 'YES',
     'Female', '~39.1', 'chr2 mild elevation — flag for QDNAseq'],
]

# ============================================================
# Metadata sheet rows
# SampleID, OriginalFilename, Batch, SeqDate, PatientID,
# PTCLSubtype, Timepoint, TreatmentStatus, CollectionDate,
# SexInferred, SexClinical, Stage, LDH, IsControl, Notes
# ============================================================

metadata_rows = [
    ['S22', 'LB_20241212_1_SI.bam', 'B3', 'Dec 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S23', 'LB_20241212_2_SI.bam', 'B3', 'Dec 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S24', 'LB_20241212_3_SI.bam', 'B3', 'Dec 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S25', 'LB_20241212_4_SI.bam', 'B3', 'Dec 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Female', '[?]', '[?]', '[?]', '[?]', None],
    ['S26', 'LB_20241212_6_SI.bam', 'B3', 'Dec 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S27', 'LB_241114_1_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S28', 'LB_241114_2_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Female', '[?]', '[?]', '[?]', '[?]', None],
    ['S29', 'LB_241114_3_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Female', '[?]', '[?]', '[?]', '[?]', None],
    ['S30', 'LB_241114_4_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Female', '[?]', '[?]', '[?]', '[?]', None],
    ['S31', 'LB_241114_5_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S32', 'LB_241114_6_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S33', 'LB_241114_7_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Male', '[?]', '[?]', '[?]', '[?]', None],
    ['S34', 'LB_241114_8_SI.bam', 'B3', 'Nov 2024', '[?]', '[?]', '[?]', '[?]', '[?]', 'Female', '[?]', '[?]', '[?]', '[?]', None],
]

# ============================================================
# Coverage_Summary sheet rows
# SampleID, Batch, GenomeMeanDepth, chrXDepth, chrYDepth,
# SexInferred, MinChrDepth, MaxChrDepth, QCFlag
# ============================================================

coverage_rows = [
    ['S22', 'B3', '1.084x', '0.553', '0.428', 'Male', '0.935 (chr22)', '1.488 (chr2)', 'PASS'],
    ['S23', 'B3', '1.084x', '0.535', '0.448', 'Male', '0.897 (chr14)', '1.205 (chr2)', 'PASS'],
    ['S24', 'B3', '1.084x', '0.520', '0.522', 'Male', '0.879 (chr13)', '1.234 (chr2)', 'PASS'],
    ['S25', 'B3', '1.084x', '1.161', '0.076', 'Female', '0.993 (chr14)', '1.380 (chr2)', 'PASS'],
    ['S26', 'B3', '1.084x', '0.521', '0.427', 'Male', '0.889 (chr15)', '1.174 (chr20)', 'FLAG: chr13/14/15/22 borderline low'],
    ['S27', 'B3', '1.390x', '0.653', '0.510', 'Male', '1.083 (chr14)', '1.678 (chr5)', 'FLAG: chr2 and chr5 elevated'],
    ['S28', 'B3', '1.770x', '1.426', '0.309', 'Female', '1.329 (chr22)', '2.052 (chr2)', 'FLAG: chr2 elevated'],
    ['S29', 'B3', '1.590x', '1.049', '0.097', 'Female', '1.217 (chr12)', '1.940 (chr2)', 'HIGH: chr6/10/12 low chr18 high'],
    ['S30', 'B3', '1.540x', '0.971', '0.089', 'Female', '1.080 (chr13)', '2.241 (chr18)', 'HIGH: chr18 +49% above mean'],
    ['S31', 'B3', '1.430x', '0.709', '0.224', 'Male', '1.145 (chr22)', '1.705 (chr2)', 'FLAG: chr2 mild elevation'],
    ['S32', 'B3', '1.680x', '0.786', '0.697', 'Male', '1.302 (chr13)', '2.062 (chr19)', 'HIGH: chr17/19/20 elevated'],
    ['S33', 'B3', '1.480x', '0.828', '0.565', 'Male', '1.187 (chr13)', '1.727 (chr19)', 'FLAG: chr19/20 mild elevation'],
    ['S34', 'B3', '1.580x', '1.398', '0.093', 'Female', '1.239 (chr14)', '1.804 (chr2)', 'FLAG: chr2 mild elevation'],
]

# ============================================================
# Load workbook and append rows to each sheet
# ============================================================

print("Loading workbook...")
wb = openpyxl.load_workbook(EXCEL_PATH)

# --- QC_Summary sheet ---
ws = wb['QC_Summary']
print(f"QC_Summary: currently {ws.max_row} rows — appending {len(qc_summary_rows)} rows")
for row_data in qc_summary_rows:
    ws.append(row_data)
print(f"QC_Summary: now {ws.max_row} rows")

# --- Metadata sheet ---
ws = wb['Metadata']
print(f"Metadata: currently {ws.max_row} rows — appending {len(metadata_rows)} rows")
for row_data in metadata_rows:
    ws.append(row_data)
print(f"Metadata: now {ws.max_row} rows")

# --- Coverage_Summary sheet ---
ws = wb['Coverage_Summary']
print(f"Coverage_Summary: currently {ws.max_row} rows — appending {len(coverage_rows)} rows")
for row_data in coverage_rows:
    ws.append(row_data)
print(f"Coverage_Summary: now {ws.max_row} rows")

# --- Save as v3 ---
print(f"Saving to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print("Done. PTCL_ctDNA_QC_Master_v3.xlsx saved successfully.")
